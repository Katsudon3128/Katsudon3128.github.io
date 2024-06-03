import openpyxl
import csv
import os
import datetime

wb = openpyxl.Workbook()

def transEmgDataToXlsx(fileName):
    file = open(fileName)
    reader = csv.reader(file)
    data_list = list(reader)
    file.close()

    sheetName = 'EMG_BT' + datetime.date.today().strftime('%y%m%d')
    wb.create_sheet(sheetName)
    resultSheet = wb[sheetName]
    # 寫入欄位名稱
    resultSheet.cell(1,1,'ID')
    resultSheet.cell(1,2,'BROKER')
    resultSheet.cell(1,3,'PRICE')
    resultSheet.cell(1,4,'BUY')
    resultSheet.cell(1,5,'SELL')

    n=1
    for i in range(5,len(data_list)-3):
        n+=1
        resultSheet.cell(n,1,data_list[i][1]).data_type = 'int'
        resultSheet.cell(n,2,data_list[i][3])
        resultSheet.cell(n,3,data_list[i][4].replace(' ','').replace('-','0')).data_type = 'int'
        resultSheet.cell(n,4,data_list[i][5].replace(' ','').replace('-','0').replace(',','')).data_type = 'int'
        resultSheet.cell(n,5,data_list[i][6].replace(' ','').replace('-','0').replace(',','')).data_type = 'int'

def transCsvDataToXlsx(fileName):
    file = open(fileName)
    reader = csv.reader(file)
    data_list = list(reader)
    file.close()

    stockId = data_list[1][1].replace('=','').replace('"','')
    sheetName = stockId + '_BT' + datetime.date.today().strftime('%y%m%d')
    wb.create_sheet(sheetName)
    resultSheet = wb[sheetName]

    # 寫入欄位名稱
    resultSheet.cell(1,1,'ID')
    resultSheet.cell(1,2,'BROKER')
    resultSheet.cell(1,3,'PRICE')
    resultSheet.cell(1,4,'BUY')
    resultSheet.cell(1,5,'SELL')

    n=1
    for i in range(3,len(data_list)):
        n+=1
        resultSheet.cell(n,1,data_list[i][0]).data_type = 'int'
        resultSheet.cell(n,2,data_list[i][1])
        resultSheet.cell(n,3,data_list[i][2]).data_type = 'int'
        resultSheet.cell(n,4,data_list[i][3].replace(',','')).data_type = 'int'
        resultSheet.cell(n,5,data_list[i][4].replace(',','')).data_type = 'int'
        
        # 如果這列有兩筆資料，就寫入
        if len(data_list[i]) == 11:
            n+=1
            resultSheet.cell(n,1,data_list[i][6]).data_type = 'int'
            resultSheet.cell(n,2,data_list[i][7])
            resultSheet.cell(n,3,data_list[i][8]).data_type = 'int'
            resultSheet.cell(n,4,data_list[i][9].replace(',','')).data_type = 'int'
            resultSheet.cell(n,5,data_list[i][10].replace(',','')).data_type = 'int'

#轉檔
dir_path = os.path.dirname(os.path.realpath(__file__))
dir = os.listdir(dir_path)

for file in dir:
    print(file)
    if file.split('.')[0] == 'EMdss004':
        transEmgDataToXlsx(dir_path + '\\' + file)
    elif file.split('.')[1].upper() == 'CSV':
        transCsvDataToXlsx(dir_path + '\\' + file)

wb.save(dir_path + '\\' + 'result.xlsx')
