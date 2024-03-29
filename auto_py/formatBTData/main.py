import openpyxl
import csv
import os

wb = openpyxl.Workbook()

# 定義function，傳入csv路徑，轉檔至xlsx
def transCsvDataToXlsx(fileName):
    # 讀取並解析csv檔
    file = open(fileName)
    reader = csv.reader(file)
    data_list = list(reader)
    file.close()

    # 抓股票代號，存成一個工作表
    stockId = data_list[1][1].replace('=','').replace('"','')
    wb.create_sheet(stockId)
    resultSheet = wb[stockId]

    # 抓取並寫入欄位名稱
    resultSheet.cell(1,1,data_list[2][0])
    resultSheet.cell(1,2,data_list[2][1])
    resultSheet.cell(1,3,data_list[2][2])
    resultSheet.cell(1,4,data_list[2][3])
    resultSheet.cell(1,5,data_list[2][4])

    n=1 # n記錄寫到第幾列
    for i in range(3,len(data_list)):
        n+=1
        resultSheet.cell(n,1,data_list[i][0]).data_type = 'int'
        resultSheet.cell(n,2,data_list[i][1])
        resultSheet.cell(n,3,data_list[i][2]).data_type = 'int'
        resultSheet.cell(n,4,data_list[i][3].replace(',','')).data_type = 'int'
        resultSheet.cell(n,5,data_list[i][4].replace(',','')).data_type = 'int'
        
        # 如果這列有兩筆資料，就寫入
        if len(data_list[i]) == 11:
            n+=1
            resultSheet.cell(n,1,data_list[i][6]).data_type = 'int'
            resultSheet.cell(n,2,data_list[i][7])
            resultSheet.cell(n,3,data_list[i][8]).data_type = 'int'
            resultSheet.cell(n,4,data_list[i][9].replace(',','')).data_type = 'int'
            resultSheet.cell(n,5,data_list[i][10].replace(',','')).data_type = 'int'

#轉檔
dir_path = os.path.dirname(os.path.realpath(__file__)) # 抓主程式目錄
dir = os.listdir(dir_path) # 抓目錄下檔案

# 如果是csv檔就呼叫function轉檔
for file in dir:
    print(file)
    if file.split('.')[1].upper() == 'CSV':
        transCsvDataToXlsx(dir_path + '\\' + file)

# 結果存至result.xlsx
wb.save(dir_path + '\\' + 'result.xlsx')
